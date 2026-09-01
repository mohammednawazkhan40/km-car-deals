"""Phase 1 tests: Excel export and CRM import."""

from pathlib import Path

from openpyxl import Workbook, load_workbook

from km_car_deals.models.enums import VehicleStatus
from km_car_deals.schemas.vehicle import VehicleCreate
from km_car_deals.services import crm, inventory
from km_car_deals.services.excel_service import ExcelService, HEADER_ALIASES

assert crm  # keep import used
assert HEADER_ALIASES


def test_export_inventory_creates_workbook(db, tmp_storage, monkeypatch):
    v = inventory.create_vehicle(
        db,
        VehicleCreate(
            manufacturer="Hyundai", model="Creta", variant="SX",
            manufacturing_year=2022, fuel_type="DIESEL",
            selling_price=1250000, mileage_km=20000, location="Mumbai",
        ),
    )
    inventory.update_status(db, v.vehicle_id, VehicleStatus.AVAILABLE.value)
    db.flush()

    svc = ExcelService(db)
    out = svc.export_inventory()

    path = out  # uses settings.EXPORT_DIR by default
    assert Path(path).exists()
    wb = load_workbook(path)
    assert "PUBLIC INVENTORY" in wb.sheetnames
    assert "INTERNAL DATA" in wb.sheetnames
    assert "CUSTOMER CRM" in wb.sheetnames
    ws = wb["PUBLIC INVENTORY"]
    # header row present
    assert ws.cell(1, 1).value == "Sl No"
    # data row has vehicle name
    values = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    assert "Creta" in str(values)


def test_header_mapping_aliases():
    assert _map("Customer Name") == "customer_name"
    assert _map("WhatsApp") == "whatsapp_number"
    assert _map("Phone Number") == "phone_number"
    assert _map("Remark") == "notes"


def _map(header):
    svc = ExcelService.__new__(ExcelService)  # no init needed
    return svc._map_header(header)


def test_import_crm_creates_and_dedupes(db, tmp_path):
    path = tmp_path / "crm.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Customer Name", "Phone", "WhatsApp", "Email", "Notes"])
    ws.append(["Ramesh", "9811111111", "9911111111", "r@x.com", "top buyer"])
    ws.append(["Ramesh", "9811111111", "9911111111", "r@x.com", "repeat"])  # duplicate
    wb.save(str(path))

    svc = ExcelService(db)
    result = svc.import_crm_excel(str(path))
    assert result["created"] == 1  # only one new customer
    assert result["updated"] == 1  # the other one matched an existing customer
    assert result["missing"] == 0
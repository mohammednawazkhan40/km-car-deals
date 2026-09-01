"""Excel inventory management using openpyxl.

Export:
  - Sheet 1 PUBLIC INVENTORY (safe fields only)
  - Sheet 2 INTERNAL DATA (includes internal info)
  - Sheet 3 CUSTOMER CRM (safe customer info)
  Includes filters, freeze panes, column widths, formatting, image thumbnails
  and hyperlinks. Sensitive info never appears in public exports.

Import:
  - Map arbitrary headers to canonical fields
  - Detect duplicates (customers / phones / vehicles) - never overwrite
    without confirmation.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy import select
from sqlalchemy.orm import Session

from km_car_deals.core.config import settings
from km_car_deals.core.logging import get_logger
from km_car_deals.models.customer import Customer
from km_car_deals.models.vehicle import Vehicle

logger = get_logger(__name__)

PUBLIC_COLUMNS = [
    "Sl No", "Stock ID", "Vehicle", "Variant", "Year", "Fuel", "Transmission",
    "Color", "Mileage", "Price", "Owner", "Status", "Location",
    "Primary Photo", "Photo 2", "Photo 3", "Photo 4", "Photo 5", "Photo 6",
]

INTERNAL_COLUMNS = [
    "Stock ID", "Registration Number", "Engine Number", "Chassis Number",
    "Purchase Price", "Selling Price", "Insurance Valid Until", "PUC Valid Until",
    "Service History", "Lead Source", "Status", "Owner Name",
]

CRM_COLUMNS = [
    "Customer ID", "Customer Name", "WhatsApp", "Phone", "Email",
    "Interested Vehicle", "Lead Status", "Last Contact", "Next Follow-up", "Notes",
]

# Header fuzzy-mapping for CRM import (case-insensitive, with underscore/space
# normalization).
HEADER_ALIASES: Dict[str, List[str]] = {
    "customer_name": ["customer name", "client", "buyer name", "name", "client name", "customer"],
    "phone_number": ["phone", "phone number", "mobile", "mobile number", "contact", "contact number", "tel"],
    "whatsapp_number": ["whatsapp", "whatsapp number", "wa number"],
    "email": ["email", "email id", "e-mail", "mail"],
    "location": ["location", "city", "state", "address"],
    "source": ["source", "lead source", "origin"],
    "lead_status": ["lead status", "status"],
    "preferred_vehicle": ["interested vehicle", "vehicle", "car", "preferred vehicle", "interested car"],
    "budget_min": ["budget min", "min budget", "min price"],
    "budget_max": ["budget max", "max budget", "max price"],
    "notes": ["notes", "remark", "remarks", "comments", "note"],
}


class ExcelService:
    def __init__(self, db: Session):
        self.db = db

    # ---------------- Export ----------------
    def export_inventory(self) -> str:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # ---- Sheet 1: PUBLIC INVENTORY ----
        ws = wb.active
        ws.title = "PUBLIC INVENTORY"
        active_vehicles = self._active_vehicles_for_export()
        ws.append(PUBLIC_COLUMNS)
        self._style_header(ws)
        self._freeze_and_filter(ws, len(PUBLIC_COLUMNS))
        sl = 1
        for v in active_vehicles:
            ws.append(self._public_row(v, sl))
            sl += 1
        self._set_widths(ws, PUBLIC_COLUMNS, "public")
        # Embed image thumbnails in a dedicated block below the table.
        self._embed_thumbnails(ws, active_vehicles, start_row=ws.max_row + 2)

        # ---- Sheet 2: INTERNAL DATA ----
        ws2 = wb.create_sheet("INTERNAL DATA")
        ws2.append(INTERNAL_COLUMNS)
        self._style_header(ws2)
        self._freeze_and_filter(ws2, len(INTERNAL_COLUMNS))
        for v in active_vehicles:
            ws2.append(self._internal_row(v))
        self._set_widths(ws2, INTERNAL_COLUMNS, "internal")

        # ---- Sheet 3: CUSTOMER CRM ----
        ws3 = wb.create_sheet("CUSTOMER CRM")
        ws3.append(CRM_COLUMNS)
        self._style_header(ws3)
        self._freeze_and_filter(ws3, len(CRM_COLUMNS))
        customers = list(self.db.execute(select(Customer).order_by(Customer.created_at)).scalars())
        for c in customers:
            ws3.append(self._crm_row(c))
        self._set_widths(ws3, CRM_COLUMNS, "crm")

        # Embed thumbnails in a dedicated grid on Sheet 1 area below the table.
        # (Single call — done only once above.)

        out_dir = Path(settings.EXPORT_DIR)
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / "KM_Car_Deals_Inventory.xlsx"
        wb.save(str(out_path))
        logger.info("Exported inventory workbook to %s", out_path)
        return str(out_path)

    def _active_vehicles_for_export(self) -> List[Vehicle]:
        return list(
            self.db.execute(
                select(Vehicle)
                .where(Vehicle.status.in_(["AVAILABLE", "READY_FOR_SALE", "NEGOTIATION", "RESERVED"]))
                .order_by(Vehicle.created_at)
            ).scalars()
        )

    def _public_row(self, v: Vehicle, sl: int) -> list:
        photos = self._photo_paths(v)
        return [
            sl,
            v.stock_id,
            v.vehicle_name or f"{v.manufacturer} {v.model}".strip(),
            v.variant,
            v.manufacturing_year,
            v.fuel_type,
            v.transmission,
            v.vehicle_color,
            v.mileage_km,
            int(v.selling_price) if v.selling_price else None,
            v.owner_count,
            v.status,
            v.location,
        ] + [p for p in photos[:6]]

    def _photo_paths(self, v: Vehicle) -> List[str]:
        photos = sorted(v.photos, key=lambda p: (not p.is_primary, p.sort_order))
        out = []
        for p in photos:
            if p.variant in ("web", "processed"):
                out.append(p.file_path)
        return out

    def _embed_thumbnails(self, ws, vehicles: List[Vehicle], start_row: int):
        """Embed small image thumbnails in a titled grid below the data."""
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
        from PIL import Image as PILImage

        ws.cell(row=start_row, column=1, value="PHOTO THUMBNAILS")
        ws.cell(row=start_row, column=1).font = Font(bold=True)
        row_idx = start_row + 1
        col = 1
        for v in vehicles:
            for path in self._photo_paths(v)[:6]:
                thumb = self._make_thumbnail(path)
                if thumb is None:
                    continue
                try:
                    img = XLImage(thumb)
                    img.width = 80
                    img.height = 60
                    cell = get_column_letter(col) + str(row_idx)
                    ws.add_image(img, cell)
                    ws.column_dimensions[get_column_letter(col)].width = 16
                    ws.row_dimensions[row_idx].height = 70
                except Exception:
                    logger.debug("Could not embed thumbnail for %s", path)
                    continue
                col += 1
        logger.info("Embedded %d vehicle thumbnails into workbook", col - 1)

    def _make_thumbnail(self, path: str) -> Optional[str]:
        """Downscale an image to a small temp thumbnail for embedding."""
        try:
            from PIL import Image as PILImage
            import tempfile, os

            if not Path(path).exists():
                return None
            img = PILImage.open(path)
            img.thumbnail((160, 120))
            fd, tmp = tempfile.mkstemp(suffix=".png")
            os.close(fd)
            img.convert("RGB").save(tmp)
            return tmp
        except Exception:
            return None

    def _internal_row(self, v: Vehicle) -> list:
        return [
            v.stock_id, v.registration_number, v.engine_number, v.chassis_number,
            v.purchase_price, v.selling_price, v.insurance_valid_until,
            v.puc_valid_until, v.service_history, "WHATSAPP", v.status, v.owner_name,
        ]

    def _crm_row(self, c: Customer) -> list:
        interests = c.interests
        vehicle_names = []
        for i in interests:
            veh = self.db.get(Vehicle, i.vehicle_id)
            if veh:
                vehicle_names.append(veh.vehicle_name or veh.model)
        next_followup = ",".join(
            str(f.scheduled_for) for f in c.followups if f.status == "PENDING"
        )
        return [
            c.customer_id, c.name, c.whatsapp_number, c.phone_number, c.email,
            ", ".join(vehicle_names) or c.preferred_vehicle,
            c.lead_status,
            c.last_inbound_at or c.last_outbound_at,
            next_followup,
            c.notes,
        ]

    def _style_header(self, ws):
        from openpyxl.styles import Font, PatternFill

        fill = PatternFill("solid", fgColor="1F4E78")
        font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font

    def _freeze_and_filter(self, ws, ncols):
        from openpyxl.utils import get_column_letter

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"

    def _set_widths(self, ws, columns: List[str], kind: str):
        from openpyxl.utils import get_column_letter

        for i in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18

    # ---------------- Import ----------------
    def import_crm_excel(
        self, file_path: str, *, confirm_duplicates: bool = False
    ) -> Dict[str, Any]:
        """Import customers from an Excel file with header mapping."""
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h else "" for h in next(rows, [])]
        mapping = {i: self._map_header(h) for i, h in enumerate(headers)}

        created, updated, skipped_duplicates, missing = [], [], [], []
        for row in rows:
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            data = self._row_to_data(row, mapping)
            if not any([data.get("name"), data.get("phone_number"), data.get("whatsapp_number"), data.get("email")]):
                missing.append("Row with no identifying info")
                continue
            customer, is_new = self._upsert_imported_customer(data, confirm_duplicates)
            if is_new:
                created.append(customer.customer_id)
            else:
                updated.append(customer.customer_id)
        return {
            "created": len(created), "updated": len(updated),
            "skipped_duplicate": len(skipped_duplicates), "missing": len(missing),
        }

    def _map_header(self, header: str) -> Optional[str]:
        norm = " ".join(header.lower().split())
        for canonical, aliases in HEADER_ALIASES.items():
            if norm in aliases or norm == canonical.replace("_", " "):
                return canonical
        return None

    def _row_to_data(self, row, mapping: Dict[int, Optional[str]]) -> Dict[str, Any]:
        data: Dict[str, Any] = {}
        for idx, canonical in mapping.items():
            if canonical is None or idx >= len(row):
                continue
            val = row[idx]
            if val is None or str(val).strip() == "":
                continue
            data[canonical] = str(val).strip()
        return data

    def _upsert_imported_customer(self, data: dict, confirm_duplicates: bool) -> Tuple[Customer, bool]:
        from km_car_deals.services import crm

        duplicate = crm.find_duplicate_customer(
            self.db,
            phone=data.get("phone_number"),
            whatsapp=data.get("whatsapp_number"),
            email=data.get("email"),
        )
        if duplicate and not confirm_duplicates:
            # Duplicate detected; do not silently overwrite.
            return duplicate, False
        payload = {
            "name": data.get("name"),
            "phone_number": data.get("phone_number"),
            "whatsapp_number": data.get("whatsapp_number"),
            "email": data.get("email"),
            "location": data.get("location"),
            "source": data.get("source") or "EXCEL_IMPORT",
            "lead_status": (data.get("lead_status") or "NEW").upper(),
        }
        customer, created = crm.upsert_customer(self.db, payload)
        if data.get("preferred_vehicle"):
            customer.preferred_vehicle = data["preferred_vehicle"]
        self.db.flush()
        return customer, created
